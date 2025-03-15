#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Power System FlowMatrix Mapper

This script processes all sheets in the system_data_all_configs.xlsx file
and generates corresponding output sheets based on the FlowMatrixTemplate.xlsx.

It handles various battery configurations, correctly maps power flows, and
places all values in their appropriate cells in the template.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import re
import os
import sys


def parse_battery_config(sheet_name):
    """
    Parse battery configuration from sheet name
    
    Args:
        sheet_name (str): Name of the sheet
        
    Returns:
        dict: Dictionary with battery configuration details
    """
    battery_nodes = []
    
    if sheet_name == "Base":
        # Base case has no batteries
        return {"count": 0, "nodes": []}
    elif len(sheet_name) == 2:
        # Single battery case: "XX"
        try:
            battery_node = int(sheet_name)
            battery_nodes.append(battery_node)
        except ValueError:
            pass
    elif len(sheet_name) == 4:
        # Two batteries case: "XXYY"
        try:
            first_battery = int(sheet_name[:2])
            second_battery = int(sheet_name[2:])
            
            battery_nodes.append(first_battery)
            battery_nodes.append(second_battery)
        except ValueError:
            pass
    
    return {
        "count": len(battery_nodes),
        "nodes": battery_nodes
    }


def find_section_ranges(data_frame):
    """
    Find the starting rows for different data sections
    
    Args:
        data_frame (DataFrame): Excel sheet data
        
    Returns:
        dict: Row indices for each section
    """
    section_ranges = {
        "branch_flow": {"start": -1, "end": -1},
        "load_data": {"start": -1, "end": -1},
        "generator_data": {"start": -1, "end": -1}
    }
    
    # Check first column for section headers
    for i, value in enumerate(data_frame.iloc[:, 0]):
        if isinstance(value, str):
            if "BRANCH POWER FLOW DATA" in value:
                section_ranges["branch_flow"]["start"] = i
            elif "LOAD DATA" in value:
                section_ranges["load_data"]["start"] = i
                if section_ranges["branch_flow"]["end"] == -1 and section_ranges["branch_flow"]["start"] != -1:
                    section_ranges["branch_flow"]["end"] = i - 1
            elif "GENERATOR DATA" in value:
                section_ranges["generator_data"]["start"] = i
                if section_ranges["load_data"]["end"] == -1 and section_ranges["load_data"]["start"] != -1:
                    section_ranges["load_data"]["end"] = i - 1
    
    # Set end of last section
    if section_ranges["generator_data"]["end"] == -1 and section_ranges["generator_data"]["start"] != -1:
        section_ranges["generator_data"]["end"] = len(data_frame)
    
    return section_ranges


def parse_system_data(sheet_data, battery_config):
    """
    Parse system data sheet to extract relevant information
    
    Args:
        sheet_data (DataFrame): Data from a system configuration sheet
        battery_config (dict): Battery configuration details
        
    Returns:
        dict: Parsed data with branch flows, generators, loads, and batteries
    """
    # Initialize with empty lists for each section
    parsed_data = {
        "branch_flow": [],
        "generators": [],
        "loads": [],
        "batteries": [],
        "slack": None
    }
    
    # Find section ranges
    sections = find_section_ranges(sheet_data)
    
    # Extract branch flow data
    if sections["branch_flow"]["start"] != -1:
        branch_header_row = sections["branch_flow"]["start"] + 1
        branch_header = sheet_data.iloc[branch_header_row].tolist()
        
        # Get data rows (starting 2 rows after header)
        branch_data_start = branch_header_row + 1
        branch_data_end = sections["branch_flow"]["end"] if sections["branch_flow"]["end"] != -1 else len(sheet_data)
        
        for i in range(branch_data_start, branch_data_end):
            row = sheet_data.iloc[i].tolist()
            
            # Check if row has valid data
            if pd.isna(row[0]) or not isinstance(row[0], (int, float)):
                continue
                
            parsed_data["branch_flow"].append({
                "fromBus": int(row[0]),
                "toBus": int(row[1]),
                "pFromMW": row[2],
                "qFromMVar": row[3],
                "pToMW": row[4],
                "qToMVar": row[5],
                "pLossMW": row[6],
                "qLossMVar": row[7],
                "loading": row[8],
                "tapRatio": row[9]
            })
    
    # Extract load data
    if sections["load_data"]["start"] != -1:
        load_header_row = sections["load_data"]["start"] + 1
        load_header = sheet_data.iloc[load_header_row].tolist()
        
        # Get data rows (starting 2 rows after header)
        load_data_start = load_header_row + 1
        load_data_end = sections["load_data"]["end"] if sections["load_data"]["end"] != -1 else len(sheet_data)
        
        for i in range(load_data_start, load_data_end):
            row = sheet_data.iloc[i].tolist()
            
            # Check if row has valid data
            if pd.isna(row[0]) or not isinstance(row[0], (int, float)):
                continue
                
            parsed_data["loads"].append({
                "bus": int(row[0]),
                "pMW": row[1],
                "qMVar": row[2]
            })
    
    # Extract generator data
    if sections["generator_data"]["start"] != -1:
        gen_header_row = sections["generator_data"]["start"] + 1
        gen_header = sheet_data.iloc[gen_header_row].tolist()
        
        # Get data rows (starting 2 rows after header)
        gen_data_start = gen_header_row + 1
        gen_data_end = sections["generator_data"]["end"] if sections["generator_data"]["end"] != -1 else len(sheet_data)
        
        for i in range(gen_data_start, gen_data_end):
            row = sheet_data.iloc[i].tolist()
            
            # Check if row has valid data
            if pd.isna(row[0]) or not isinstance(row[0], (int, float)):
                continue
                
            gen_data = {
                "bus": int(row[0]),
                "type": row[1],
                "pMW": row[2],
                "qMVar": row[3],
                "vset": row[4],
                "isBattery": row[5] == "Yes" if len(row) > 5 else False
            }
            
            if gen_data["isBattery"]:
                parsed_data["batteries"].append(gen_data)
            elif gen_data["type"] == "Slack":
                parsed_data["slack"] = gen_data
            else:
                parsed_data["generators"].append(gen_data)
    
    return parsed_data


def fill_generators(template_ws, data):
    """
    Fill generator cells in the template
    
    Args:
        template_ws (Worksheet): Template worksheet
        data (dict): Parsed system data
        
    Returns:
        dict: Mapping of generator cells and values
    """
    generator_mapping = {}
    
    # Map generators in the header row (row 2)
    header_generator_cells = [
        {"ref": "D2", "genType": "Slack"},               # P Slack
        {"ref": "E2", "genType": "PV", "genBus": 2},    # P Gen 2
        {"ref": "F2", "genType": "PV", "genBus": 3},    # P Gen 3
        {"ref": "G2", "genType": "PV", "genBus": 6},    # P Gen 6
        {"ref": "H2", "genType": "PV", "genBus": 8}     # P Gen 8
    ]
    
    # Battery cells in header row - set to 0 if no data
    header_battery_cells = [
        {"ref": "I2"},  # P Battery 1
        {"ref": "J2"}   # P Battery 2
    ]
    
    # Add generators in the diagonal (main power flow matrix)
    diagonal_generator_cells = [
        {"ref": "K3", "genType": "Slack"},               # P Slack
        {"ref": "L4", "genType": "PV", "genBus": 2},    # P Gen 2
        {"ref": "M5", "genType": "PV", "genBus": 3},    # P Gen 3
        {"ref": "P6", "genType": "PV", "genBus": 6},    # P Gen 6
        {"ref": "R7", "genType": "PV", "genBus": 8}     # P Gen 8
    ]
    
    # Process each generator cell
    for cell in header_generator_cells + diagonal_generator_cells:
        gen_value = None
        
        if cell["genType"] == "Slack" and data["slack"]:
            gen_value = data["slack"]["pMW"]
        elif cell["genType"] == "PV":
            gen = next((g for g in data["generators"] if g["bus"] == cell["genBus"]), None)
            if gen:
                gen_value = gen["pMW"]
        
        if gen_value is not None:
            template_ws[cell["ref"]] = gen_value
            generator_mapping[cell["ref"]] = gen_value
        else:
            template_ws[cell["ref"]] = 0
            generator_mapping[cell["ref"]] = 0
    
    # Set battery cells to 0 if no data
    for cell in header_battery_cells:
        if cell["ref"] not in generator_mapping:
            template_ws[cell["ref"]] = 0
            generator_mapping[cell["ref"]] = 0
    
    return generator_mapping


def fill_flows(template_ws, data):
    """
    Fill power flow cells in the template
    
    Args:
        template_ws (Worksheet): Template worksheet
        data (dict): Parsed system data
        
    Returns:
        dict: Mapping of flow cells and values
    """
    flow_mapping = {
        "from": {},
        "to": {}
    }
    
    # Define flow "From" cells with their corresponding bus numbers
    flow_from_cells = [
        {"ref": "L10", "fromBus": 1, "toBus": 2},   # From 12
        {"ref": "O10", "fromBus": 1, "toBus": 5},   # From 15
        {"ref": "M11", "fromBus": 2, "toBus": 3},   # From 23
        {"ref": "N11", "fromBus": 2, "toBus": 4},   # From 24
        {"ref": "O11", "fromBus": 2, "toBus": 5},   # From 25
        {"ref": "N12", "fromBus": 3, "toBus": 4},   # From 34
        {"ref": "O13", "fromBus": 4, "toBus": 5},   # From 45
        {"ref": "Q13", "fromBus": 4, "toBus": 7},   # From 47
        {"ref": "S13", "fromBus": 4, "toBus": 9},   # From 49
        {"ref": "P14", "fromBus": 6, "toBus": 5},   # From 65 (used when flow is negative)
        {"ref": "U15", "fromBus": 6, "toBus": 11},  # From 611
        {"ref": "V15", "fromBus": 6, "toBus": 12},  # From 612
        {"ref": "W15", "fromBus": 6, "toBus": 13},  # From 613
        {"ref": "R16", "fromBus": 8, "toBus": 7},   # From 87 (used when flow is negative)
        {"ref": "S16", "fromBus": 7, "toBus": 9},   # From 79
        {"ref": "T18", "fromBus": 9, "toBus": 10},  # From 910
        {"ref": "X18", "fromBus": 9, "toBus": 14},  # From 914
        {"ref": "U19", "fromBus": 10, "toBus": 11}, # From 1011
        {"ref": "W21", "fromBus": 12, "toBus": 13}, # From 1213
        {"ref": "X22", "fromBus": 13, "toBus": 14}  # From 1314
    ]
    
    # Define flow "To" cells with their corresponding bus numbers
    flow_to_cells = [
        {"ref": "K11", "fromBus": 1, "toBus": 2},   # To 12
        {"ref": "L12", "fromBus": 2, "toBus": 3},   # To 23
        {"ref": "L13", "fromBus": 2, "toBus": 4},   # To 24
        {"ref": "M13", "fromBus": 3, "toBus": 4},   # To 34
        {"ref": "K14", "fromBus": 1, "toBus": 5},   # To 15
        {"ref": "L14", "fromBus": 2, "toBus": 5},   # To 25
        {"ref": "N14", "fromBus": 4, "toBus": 5},   # To 45
        {"ref": "O15", "fromBus": 6, "toBus": 5},   # To 65 (used when flow is positive)
        {"ref": "N16", "fromBus": 4, "toBus": 7},   # To 47
        {"ref": "Q17", "fromBus": 8, "toBus": 7},   # To 87 (used when flow is positive)
        {"ref": "N18", "fromBus": 4, "toBus": 9},   # To 49
        {"ref": "Q18", "fromBus": 7, "toBus": 9},   # To 79
        {"ref": "S19", "fromBus": 9, "toBus": 10},  # To 910
        {"ref": "P20", "fromBus": 6, "toBus": 11},  # To 611
        {"ref": "T20", "fromBus": 10, "toBus": 11}, # To 1011
        {"ref": "P21", "fromBus": 6, "toBus": 12},  # To 612
        {"ref": "P22", "fromBus": 6, "toBus": 13},  # To 613
        {"ref": "V22", "fromBus": 12, "toBus": 13}, # To 1213
        {"ref": "S23", "fromBus": 9, "toBus": 14},  # To 914
        {"ref": "W23", "fromBus": 13, "toBus": 14}  # To 1314
    ]
    
    # Special handling for transformer flows (65 and 87)
    def handle_transformer_flow(branch, from_bus, to_bus):
        # For 6→5 flow
        if (from_bus == 6 and to_bus == 5) or (from_bus == 5 and to_bus == 6):
            flow = branch["pFromMW"] if branch["fromBus"] == 6 else -branch["pToMW"]
            if flow > 0:
                # Positive flow goes in O15 (To cell)
                template_ws["O15"] = flow
                template_ws["P14"] = 0  # Set unused cell to 0
                flow_mapping["to"]["O15"] = flow
                flow_mapping["from"]["P14"] = 0
            else:
                # Negative flow goes in P14 (From cell) as absolute value
                template_ws["P14"] = abs(flow)
                template_ws["O15"] = 0  # Set unused cell to 0
                flow_mapping["from"]["P14"] = abs(flow)
                flow_mapping["to"]["O15"] = 0
            return True
            
        # For 8→7 flow
        if (from_bus == 8 and to_bus == 7) or (from_bus == 7 and to_bus == 8):
            flow = branch["pFromMW"] if branch["fromBus"] == 8 else -branch["pToMW"]
            if flow > 0:
                # Positive flow goes in Q17 (To cell)
                template_ws["Q17"] = flow
                template_ws["R16"] = 0  # Set unused cell to 0
                flow_mapping["to"]["Q17"] = flow
                flow_mapping["from"]["R16"] = 0
            else:
                # Negative flow goes in R16 (From cell) as absolute value
                template_ws["R16"] = abs(flow)
                template_ws["Q17"] = 0  # Set unused cell to 0
                flow_mapping["from"]["R16"] = abs(flow)
                flow_mapping["to"]["Q17"] = 0
            return True
            
        return False
    
    # Process each branch flow
    for branch in data["branch_flow"]:
        # Check if this is a transformer flow that needs special handling
        if handle_transformer_flow(branch, branch["fromBus"], branch["toBus"]):
            continue
            
        # For non-transformer flows, process as before
        for cell in flow_from_cells:
            if branch["fromBus"] == cell["fromBus"] and branch["toBus"] == cell["toBus"]:
                if branch["pFromMW"] > 0:
                    template_ws[cell["ref"]] = branch["pFromMW"]
                    flow_mapping["from"][cell["ref"]] = branch["pFromMW"]
                else:
                    template_ws[cell["ref"]] = 0
                    flow_mapping["from"][cell["ref"]] = 0
                    
        for cell in flow_to_cells:
            if branch["fromBus"] == cell["fromBus"] and branch["toBus"] == cell["toBus"]:
                if branch["pToMW"] > 0:
                    template_ws[cell["ref"]] = branch["pToMW"]
                    flow_mapping["to"][cell["ref"]] = branch["pToMW"]
                else:
                    template_ws[cell["ref"]] = 0
                    flow_mapping["to"][cell["ref"]] = 0
    
    return flow_mapping


def calculate_node_dissipation(data):
    """
    Calculate dissipation (losses) for each node
    
    Args:
        data (dict): Parsed system data
        
    Returns:
        dict: Mapping of node numbers to their dissipation values
    """
    node_dissipation = {}
    
    # Initialize all node dissipations to 0
    all_nodes = set()
    for branch in data["branch_flow"]:
        all_nodes.add(branch["fromBus"])
        all_nodes.add(branch["toBus"])
    
    for node in all_nodes:
        node_dissipation[node] = 0
    
    # Calculate dissipation for each node
    for branch in data["branch_flow"]:
        # Add half of the loss to each connected node
        from_bus = branch["fromBus"]
        to_bus = branch["toBus"]
        loss = branch["pLossMW"]
        
        # Add half of the loss to each node
        node_dissipation[from_bus] += loss / 2
        node_dissipation[to_bus] += loss / 2
    
    return node_dissipation


def fill_loads(template_ws, data):
    """
    Fill load cells in the template
    
    Args:
        template_ws (Worksheet): Template worksheet
        data (dict): Parsed system data
        
    Returns:
        dict: Mapping of load cells and values
    """
    load_mapping = {}
    
    # Define load cells with their corresponding bus numbers
    load_cells = [
        {"ref": "Y10", "bus": 1},  # Load 1
        {"ref": "Y11", "bus": 2},  # Load 2
        {"ref": "Y12", "bus": 3},  # Load 3
        {"ref": "Y13", "bus": 4},  # Load 4
        {"ref": "Y14", "bus": 5},  # Load 5
        {"ref": "Y15", "bus": 6},  # Load 6
        {"ref": "Y16", "bus": 7},  # Load 7
        {"ref": "Y17", "bus": 8},  # Load 8
        {"ref": "Y18", "bus": 9},  # Load 9
        {"ref": "Y19", "bus": 10}, # Load 10
        {"ref": "Y20", "bus": 11}, # Load 11
        {"ref": "Y21", "bus": 12}, # Load 12
        {"ref": "Y22", "bus": 13}, # Load 13
        {"ref": "Y23", "bus": 14}  # Load 14
    ]
    
    # Process each load cell
    for cell in load_cells:
        load = next((l for l in data["loads"] if l["bus"] == cell["bus"]), None)
        
        if load:
            template_ws[cell["ref"]] = load["pMW"]
            load_mapping[cell["ref"]] = load["pMW"]
        else:
            # If no load found for this bus, set to 0
            template_ws[cell["ref"]] = 0
            load_mapping[cell["ref"]] = 0
    
    return load_mapping


def fill_dissipation(template_ws, data):
    """
    Fill dissipation (losses) for each node in the template
    
    Args:
        template_ws (Worksheet): Template worksheet
        data (dict): Parsed system data
        
    Returns:
        dict: Mapping of dissipation cells and values
    """
    dissipation_mapping = {}
    
    # Calculate dissipation for each node
    node_dissipation = calculate_node_dissipation(data)
    
    # Define dissipation cells with their corresponding bus numbers
    dissipation_cells = [
        {"ref": "Z10", "bus": 1},  # Dissipation 1
        {"ref": "Z11", "bus": 2},  # Dissipation 2
        {"ref": "Z12", "bus": 3},  # Dissipation 3
        {"ref": "Z13", "bus": 4},  # Dissipation 4
        {"ref": "Z14", "bus": 5},  # Dissipation 5
        {"ref": "Z15", "bus": 6},  # Dissipation 6
        {"ref": "Z16", "bus": 7},  # Dissipation 7
        {"ref": "Z17", "bus": 8},  # Dissipation 8
        {"ref": "Z18", "bus": 9},  # Dissipation 9
        {"ref": "Z19", "bus": 10}, # Dissipation 10
        {"ref": "Z20", "bus": 11}, # Dissipation 11
        {"ref": "Z21", "bus": 12}, # Dissipation 12
        {"ref": "Z22", "bus": 13}, # Dissipation 13
        {"ref": "Z23", "bus": 14}  # Dissipation 14
    ]
    
    # Process each dissipation cell
    for cell in dissipation_cells:
        bus_num = cell["bus"]
        if bus_num in node_dissipation:
            template_ws[cell["ref"]] = node_dissipation[bus_num]
            dissipation_mapping[cell["ref"]] = node_dissipation[bus_num]
        else:
            # If no dissipation found for this bus, set to 0
            template_ws[cell["ref"]] = 0
            dissipation_mapping[cell["ref"]] = 0
    
    return dissipation_mapping


def fill_batteries(template_ws, data, battery_config):
    """
    Fill battery cells in the template
    
    Args:
        template_ws (Worksheet): Template worksheet
        data (dict): Parsed system data
        battery_config (dict): Battery configuration details
        
    Returns:
        dict: Mapping of battery cells and values
    """
    battery_mapping = {}
    
    # Map for column letters based on bus numbers
    bus_to_column_map = {
        1: 'K', 2: 'L', 3: 'M', 4: 'N', 5: 'O', 
        6: 'P', 7: 'Q', 8: 'R', 9: 'S', 10: 'T',
        11: 'U', 12: 'V', 13: 'W', 14: 'X'
    }
    
    # Header row battery cells
    if data["batteries"]:
        if len(data["batteries"]) >= 1:
            battery1 = data["batteries"][0]
            # Put in header row (I2)
            template_ws["I2"] = battery1["pMW"]
            battery_mapping["I2"] = battery1["pMW"]
            
            # Put in the column corresponding to the battery bus (row 8)
            col_letter = bus_to_column_map.get(battery1["bus"])
            if col_letter:
                cell_ref = f"{col_letter}8"
                template_ws[cell_ref] = battery1["pMW"]
                battery_mapping[cell_ref] = battery1["pMW"]
        
        if len(data["batteries"]) >= 2:
            battery2 = data["batteries"][1]
            # Put in header row (J2)
            template_ws["J2"] = battery2["pMW"]
            battery_mapping["J2"] = battery2["pMW"]
            
            # Put in the column corresponding to the battery bus (row 9)
            col_letter = bus_to_column_map.get(battery2["bus"])
            if col_letter:
                cell_ref = f"{col_letter}9"
                template_ws[cell_ref] = battery2["pMW"]
                battery_mapping[cell_ref] = battery2["pMW"]
    else:
        # If no batteries, set cells to 0
        template_ws["I2"] = 0
        battery_mapping["I2"] = 0
        template_ws["J2"] = 0
        battery_mapping["J2"] = 0
    
    return battery_mapping


def fill_template_sheet(template_ws, data, battery_config):
    """
    Fill template sheet with extracted data
    
    Args:
        template_ws (Worksheet): Template worksheet
        data (dict): Parsed system data
        battery_config (dict): Battery configuration details
        
    Returns:
        dict: Mapping of all filled cells
    """
    generator_mapping = fill_generators(template_ws, data)
    flow_mapping = fill_flows(template_ws, data)
    load_mapping = fill_loads(template_ws, data)
    battery_mapping = fill_batteries(template_ws, data, battery_config)
    dissipation_mapping = fill_dissipation(template_ws, data)
    
    # Print summary of mapped cells
    print(f"Mapped {len(generator_mapping)} generator cells")
    print(f"Mapped {len(flow_mapping['from'])} flow 'from' cells")
    print(f"Mapped {len(flow_mapping['to'])} flow 'to' cells")
    print(f"Mapped {len(load_mapping)} load cells")
    print(f"Mapped {len(battery_mapping)} battery cells")
    print(f"Mapped {len(dissipation_mapping)} dissipation cells")
    
    return {
        "generators": generator_mapping,
        "flows": flow_mapping,
        "loads": load_mapping,
        "batteries": battery_mapping,
        "dissipation": dissipation_mapping
    }


def process_all_configurations(input_file, template_file, output_file):
    """
    Process all sheets in the input file and create output Excel file
    
    Args:
        input_file (str): Path to system_data_all_configs.xlsx
        template_file (str): Path to FlowMatrixTemplate.xlsx
        output_file (str): Path to output Excel file
    """
    try:
        print("Reading input files...")
        
        # Read input and template files
        all_configs_wb = load_workbook(input_file, data_only=True)
        template_wb = load_workbook(template_file)
        
        # Create output workbook
        output_wb = Workbook()
        # Remove default sheet
        if "Sheet" in output_wb.sheetnames:
            output_wb.remove(output_wb["Sheet"])
        
        # Get template worksheet
        template_ws = template_wb.active
        
        sheet_count = len(all_configs_wb.sheetnames)
        print(f"Found {sheet_count} configuration sheets")
        
        # Process each sheet in the all_configs workbook
        for sheet_name in all_configs_wb.sheetnames:
            print(f"\nProcessing configuration: {sheet_name}")
            
            # Determine battery configurations from sheet name
            battery_config = parse_battery_config(sheet_name)
            print(f"Battery configuration: {battery_config}")
            
            # Read data from the current sheet
            sheet_data = pd.DataFrame(all_configs_wb[sheet_name].values)
            
            # Parse data from the sheet
            parsed_data = parse_system_data(sheet_data, battery_config)
            
            # Create a new sheet based on the template
            new_ws = output_wb.create_sheet(title=sheet_name)
            
            # Copy template to new sheet
            for row in template_ws.iter_rows(values_only=True):
                new_ws.append(row)
            
            # Fill the template with the extracted data
            fill_template_sheet(new_ws, parsed_data, battery_config)
        
        # Save the output file
        output_wb.save(output_file)
        print(f"Output file created successfully at: {output_file}")
        
    except Exception as e:
        print(f"Error processing configurations: {str(e)}")
        raise


if __name__ == "__main__":
    # File paths
    input_file = "system_data_all_configs.xlsx"
    template_file = "data/FlowMatrixTemplate.xlsx"
    output_file = "output_flow_matrices.xlsx"
    
    # Check if files exist
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' not found")
        sys.exit(1)
    
    if not os.path.exists(template_file):
        print(f"Error: Template file '{template_file}' not found")
        sys.exit(1)
    
    # Process all configurations
    process_all_configurations(input_file, template_file, output_file)